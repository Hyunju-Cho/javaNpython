from openpyxl import Workbook #파일 생성

wb=Workbook()
ws=wb.active

ws['A1']='123123'
ws['A2']='안녕하세요'
ws.title="처음에자동sheet"
ws.sheet_properties.tabColor="cccccc"

ms1=wb.create_sheet("MySheet")
ms1['B1']='아아아'
ms1['B2']='하하하'

ms2=wb.create_sheet("MySheet",0)
ms2['C1']='test'
ms2['C2']='test'

ms3=wb["MySheet"]

wb.save('a.xlsx')
wb.close()